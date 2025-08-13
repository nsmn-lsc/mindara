"""
Vistas para la aplicación de reportes
Sistema de generación de reportes para eventos de Mindara
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views.generic import View
from django.utils import timezone
from django.db.models import Q
from datetime import datetime, timedelta
import os
import io
import locale
import pytz

# Configurar idioma español para fechas
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES')
    except:
        pass  # Si no se puede configurar español, usar por defecto

# Imports para generar archivos
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import inch
from django.conf import settings

from apps.eventos.models import Evento
from .models import ReporteGenerado


def formatear_fecha_espanol(fecha):
    """Formatear fecha en español"""
    meses = {
        'January': 'enero', 'February': 'febrero', 'March': 'marzo',
        'April': 'abril', 'May': 'mayo', 'June': 'junio',
        'July': 'julio', 'August': 'agosto', 'September': 'septiembre',
        'October': 'octubre', 'November': 'noviembre', 'December': 'diciembre'
    }
    
    fecha_str = fecha.strftime('%d de %B de %Y')
    for ingles, espanol in meses.items():
        fecha_str = fecha_str.replace(ingles, espanol)
    
    return fecha_str


def formatear_fecha_hora_espanol(fecha):
    """Formatear fecha y hora en español con zona horaria local"""
    # Convertir a zona horaria local si la fecha está en UTC
    if timezone.is_aware(fecha):
        # Obtener la zona horaria configurada en Django
        local_tz = pytz.timezone(settings.TIME_ZONE)
        fecha_local = fecha.astimezone(local_tz)
    else:
        fecha_local = fecha
    
    meses = {
        'January': 'enero', 'February': 'febrero', 'March': 'marzo',
        'April': 'abril', 'May': 'mayo', 'June': 'junio',
        'July': 'julio', 'August': 'agosto', 'September': 'septiembre',
        'October': 'octubre', 'November': 'noviembre', 'December': 'diciembre'
    }
    
    fecha_str = fecha_local.strftime('%d de %B de %Y a las %H:%M')
    for ingles, espanol in meses.items():
        fecha_str = fecha_str.replace(ingles, espanol)
    
    return fecha_str


def formatear_mes_ano_espanol(fecha):
    """Formatear mes y año en español"""
    meses = {
        'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo',
        'April': 'Abril', 'May': 'Mayo', 'June': 'Junio',
        'July': 'Julio', 'August': 'Agosto', 'September': 'Septiembre',
        'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
    }
    
    fecha_str = fecha.strftime('%B %Y')
    for ingles, espanol in meses.items():
        fecha_str = fecha_str.replace(ingles, espanol)
    
    return fecha_str


@method_decorator(login_required, name='dispatch')
class ReportesView(View):
    """
    Vista principal para el módulo de reportes
    """
    
    def get(self, request):
        """
        Mostrar la página principal de reportes
        """
        # Obtener reportes recientes del usuario
        reportes_recientes = ReporteGenerado.objects.filter(
            generado_por=request.user
        ).order_by('-fecha_generacion')[:10]
        
        context = {
            'title': 'Reportes',
            'current_section': 'reportes',
            'reportes_recientes': reportes_recientes,
        }
        return render(request, 'reportes/index.html', context)


@login_required
def generar_reporte_agenda(request):
    """Generar reporte de eventos en agenda (próximos eventos)"""
    formato = request.GET.get('formato', 'xlsx')
    incluir_detalles = request.GET.get('incluir_detalles', 'true') == 'true'
    solo_confirmados = request.GET.get('solo_confirmados', 'false') == 'true'
    
    # Obtener eventos según el nivel del usuario
    user = request.user
    if user.is_admin() or user.is_manager():
        eventos = Evento.objects.all()
    else:
        eventos = Evento.objects.filter(usuario=user)
    
    # Filtrar eventos futuros
    hoy = timezone.now().date()
    eventos = eventos.filter(fecha_evento__gte=hoy)
    
    if solo_confirmados:
        eventos = eventos.filter(etapa='confirmado')
    
    eventos = eventos.order_by('fecha_evento', 'hora_evento')
    
    # Crear registro del reporte
    reporte = ReporteGenerado.objects.create(
        tipo='agenda',
        formato=formato,
        titulo='Eventos en Agenda',
        generado_por=user,
        incluir_detalles=incluir_detalles,
        solo_confirmados=solo_confirmados,
        total_eventos=eventos.count()
    )
    
    if formato == 'xlsx':
        return _generar_excel_eventos(eventos, reporte, incluir_detalles)
    else:
        return _generar_pdf_eventos(eventos, reporte, incluir_detalles)


@login_required
def generar_reporte_semana(request):
    """Generar reporte de eventos de la semana"""
    formato = request.GET.get('formato', 'xlsx')
    incluir_detalles = request.GET.get('incluir_detalles', 'true') == 'true'
    
    # Obtener eventos según el nivel del usuario
    user = request.user
    if user.is_admin() or user.is_manager():
        eventos = Evento.objects.all()
    else:
        eventos = Evento.objects.filter(usuario=user)
    
    # Filtrar eventos de esta semana
    hoy = timezone.now().date()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    fin_semana = inicio_semana + timedelta(days=6)
    
    eventos = eventos.filter(
        fecha_evento__gte=inicio_semana,
        fecha_evento__lte=fin_semana
    ).order_by('fecha_evento', 'hora_evento')
    
    # Crear registro del reporte
    reporte = ReporteGenerado.objects.create(
        tipo='semana',
        formato=formato,
        titulo=f'Eventos de la Semana ({inicio_semana.strftime("%d/%m")} - {formatear_fecha_espanol(fin_semana)})',
        generado_por=user,
        fecha_inicio_filtro=inicio_semana,
        fecha_fin_filtro=fin_semana,
        incluir_detalles=incluir_detalles,
        total_eventos=eventos.count()
    )
    
    if formato == 'xlsx':
        return _generar_excel_eventos(eventos, reporte, incluir_detalles)
    else:
        return _generar_pdf_eventos(eventos, reporte, incluir_detalles)


@login_required
def generar_reporte_mes(request):
    """Generar reporte de eventos del mes"""
    formato = request.GET.get('formato', 'xlsx')
    incluir_detalles = request.GET.get('incluir_detalles', 'true') == 'true'
    
    # Obtener eventos según el nivel del usuario
    user = request.user
    if user.is_admin() or user.is_manager():
        eventos = Evento.objects.all()
    else:
        eventos = Evento.objects.filter(usuario=user)
    
    # Filtrar eventos de este mes
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    if hoy.month == 12:
        fin_mes = hoy.replace(year=hoy.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        fin_mes = hoy.replace(month=hoy.month + 1, day=1) - timedelta(days=1)
    
    eventos = eventos.filter(
        fecha_evento__gte=inicio_mes,
        fecha_evento__lte=fin_mes
    ).order_by('fecha_evento', 'hora_evento')
    
    # Crear registro del reporte
    reporte = ReporteGenerado.objects.create(
        tipo='mes',
        formato=formato,
        titulo=f'Eventos del Mes ({formatear_mes_ano_espanol(inicio_mes)})',
        generado_por=user,
        fecha_inicio_filtro=inicio_mes,
        fecha_fin_filtro=fin_mes,
        incluir_detalles=incluir_detalles,
        total_eventos=eventos.count()
    )
    
    if formato == 'xlsx':
        return _generar_excel_eventos(eventos, reporte, incluir_detalles)
    else:
        return _generar_pdf_eventos(eventos, reporte, incluir_detalles)


@login_required
def generar_reporte_carpeta_ejecutiva(request):
    """Generar reporte de eventos con carpeta ejecutiva"""
    formato = request.GET.get('formato', 'xlsx')
    incluir_detalles = request.GET.get('incluir_detalles', 'true') == 'true'
    
    # Obtener eventos según el nivel del usuario
    user = request.user
    if user.is_admin() or user.is_manager():
        eventos = Evento.objects.all()
    else:
        eventos = Evento.objects.filter(usuario=user)
    
    # Filtrar eventos con carpeta ejecutiva
    eventos = eventos.filter(carpeta_ejecutiva=True).order_by('fecha_evento', 'hora_evento')
    
    # Crear registro del reporte
    reporte = ReporteGenerado.objects.create(
        tipo='carpeta_ejecutiva',
        formato=formato,
        titulo='Eventos con Carpeta Ejecutiva',
        generado_por=user,
        incluir_detalles=incluir_detalles,
        total_eventos=eventos.count()
    )
    
    if formato == 'xlsx':
        return _generar_excel_eventos(eventos, reporte, incluir_detalles)
    else:
        return _generar_pdf_eventos(eventos, reporte, incluir_detalles)


def _generar_excel_eventos(eventos, reporte, incluir_detalles=True):
    """Genera un archivo Excel con los eventos"""
    
    # Crear el workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eventos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título del reporte alineado con logo
    ws.merge_cells('A1:G1')
    ws['A1'] = f"{reporte.titulo}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
    
    # Intentar agregar logo en Excel al mismo nivel que el título
    try:
        logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'img', 'logo_reportes.png')
        if os.path.exists(logo_path):
            img = ExcelImage(logo_path)
            img.width = 75  # Aproximadamente 2 cm
            img.height = 75  # Aproximadamente 2 cm
            ws.add_image(img, 'H1')  # Colocar en columna H, misma fila que el título
    except:
        pass  # Si no se puede cargar el logo, continuar sin él
    
    # Información adicional
    ws['A2'] = f"Generado por: {reporte.generado_por.get_full_name() or reporte.generado_por.username}"
    ws['A3'] = f"Fecha de generación: {formatear_fecha_hora_espanol(reporte.fecha_generacion)}"
    ws['A4'] = f"Total de eventos: {reporte.total_eventos}"
    
    # Encabezados de las columnas
    row = 6
    headers = ['Evento', 'Fecha', 'Hora', 'Sede', 'Prioridad', 'Estado', 'Responsable']
    
    if incluir_detalles:
        headers.extend(['Objetivo', 'Participantes', 'Aforo', 'Duración'])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos de los eventos
    row += 1
    for evento in eventos:
        col = 1
        ws.cell(row=row, column=col, value=evento.nombre_evento).border = border
        col += 1
        ws.cell(row=row, column=col, value=formatear_fecha_espanol(evento.fecha_evento)).border = border
        col += 1
        ws.cell(row=row, column=col, value=evento.hora_evento.strftime('%H:%M')).border = border
        col += 1
        ws.cell(row=row, column=col, value=evento.sede).border = border
        col += 1
        ws.cell(row=row, column=col, value=evento.get_prioridad_display()).border = border
        col += 1
        ws.cell(row=row, column=col, value=evento.get_etapa_display()).border = border
        col += 1
        responsable = evento.usuario.get_full_name() or evento.usuario.username
        ws.cell(row=row, column=col, value=responsable).border = border
        
        if incluir_detalles:
            col += 1
            ws.cell(row=row, column=col, value=evento.objetivo[:100] + ('...' if len(evento.objetivo) > 100 else '')).border = border
            col += 1
            ws.cell(row=row, column=col, value=evento.participantes).border = border
            col += 1
            ws.cell(row=row, column=col, value=evento.aforo).border = border
            col += 1
            ws.cell(row=row, column=col, value=f"{evento.duracion_real} hrs").border = border
        
        row += 1
    
    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar el archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear la respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{reporte.nombre_archivo}"'
    
    return response


def _generar_pdf_eventos(eventos, reporte, incluir_detalles=True):
    """Genera un archivo PDF con los eventos"""
    
    # Crear el buffer
    buffer = io.BytesIO()
    
    # Crear el documento
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Crear encabezado con título y logo al mismo nivel
    header_data = []
    
    # Crear el título
    title_text = f"{reporte.titulo}"
    title_para = Paragraph(title_text, title_style)
    
    # Intentar agregar logo al mismo nivel
    logo_element = None
    try:
        logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'img', 'logo_reportes.png')
        if os.path.exists(logo_path):
            logo_element = Image(logo_path, width=2*inch/2.54, height=2*inch/2.54)  # 2 cm cuadrados
    except:
        pass  # Si no se puede cargar el logo, continuar sin él
    
    # Crear tabla para alinear título y logo
    if logo_element:
        header_table = Table([[title_para, logo_element]], colWidths=[4.5*inch, 2*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),   # Título a la izquierda
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),  # Logo a la derecha
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Centrar verticalmente
        ]))
        elements.append(header_table)
    else:
        elements.append(title_para)
    
    elements.append(Spacer(1, 20))
    
    # Información del reporte
    info_style = styles['Normal']
    info_data = [
        f"Generado por: {reporte.generado_por.get_full_name() or reporte.generado_por.username}",
        f"Fecha de generación: {formatear_fecha_hora_espanol(reporte.fecha_generacion)}",
        f"Total de eventos: {reporte.total_eventos}"
    ]
    
    for info in info_data:
        elements.append(Paragraph(info, info_style))
    
    elements.append(Spacer(1, 20))
    
    # Crear la tabla
    data = [['Evento', 'Fecha', 'Hora', 'Sede', 'Prioridad', 'Estado', 'Responsable']]
    
    for evento in eventos:
        responsable = evento.usuario.get_full_name() or evento.usuario.username
        row = [
            evento.nombre_evento[:30] + ('...' if len(evento.nombre_evento) > 30 else ''),
            formatear_fecha_espanol(evento.fecha_evento),
            evento.hora_evento.strftime('%H:%M'),
            evento.sede[:20] + ('...' if len(evento.sede) > 20 else ''),
            evento.get_prioridad_display(),
            evento.get_etapa_display(),
            responsable[:20] + ('...' if len(responsable) > 20 else '')
        ]
        data.append(row)
    
    # Crear la tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),  # Color del header igual al Excel
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Fondo blanco para las celdas
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Construir el PDF
    doc.build(elements)
    
    # Crear la respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{reporte.nombre_archivo}"'
    
    return response


@login_required
def historial_reportes(request):
    """Vista para mostrar el historial de reportes generados"""
    reportes = ReporteGenerado.objects.filter(
        generado_por=request.user
    ).order_by('-fecha_generacion')
    
    context = {
        'title': 'Historial de Reportes',
        'current_section': 'reportes',
        'reportes': reportes,
    }
    return render(request, 'reportes/historial.html', context)
